#! /Volumes/Media/venv/myDjango_venv/bin/python3

from typing import Dict
from openpyxl import load_workbook
import os,json
from itertools import islice
from collections import OrderedDict

# loading the excel file from the same location.
xcel_file = load_workbook('Untitled.xlsx')

# Extracting the sheets names
xcel_sheet_names = xcel_file.sheetnames

# Multiple sheet reading logic
for (i, sheet) in enumerate(xcel_sheet_names):
    header_row = []
    ws = xcel_file[sheet]
    m_row = xcel_file[sheet].max_row
    m_col = xcel_file[sheet].max_column

    # print(m_row, m_col)
    # calcluating the active cell dimension.
    start_cell, end_cell = ws.calculate_dimension().split(':')
    header_row = tuple(v.value for v in (ws[1]))
    # print(len(header_row))

    cell_values = ws.iter_rows(min_row = 2,max_row = m_row, max_col = m_col,values_only = True)
    # create the dict with col#1 as the key and remaining data as child dict.
    final_dict = dict()
    header_dict = dict()
    for row in (cell_values):
        if row[0] != None:
            id_num = row[0]
            for col in range(m_col):
                if row[col] != None:
                    header_dict[header_row[col]] = row[col]
            final_dict[id_num] = header_dict

    # print(final_dict)
    # Creating the JSON file and writing the data 
    filename = 'json_file_' + sheet + '.json'
    with open (filename,'w') as f:
        json.dump(final_dict,f,indent=4)




   